from openpyxl.styles import Color, PatternFill, Border, Side, Font, Alignment
from openpyxl import Workbook, load_workbook
from datetime import datetime
from time import sleep

from .utils import logger
from .wallet import Wallet


class Excel:
    def __init__(self, total_len: int):
        workbook = Workbook()
        sheet = workbook.active
        self.file_name = f'{total_len}accs_{datetime.now().strftime("%d_%m_%Y_%H_%M_%S")}.xlsx'

        sheet['A1'] = 'EVM privatekey'
        sheet['B1'] = 'EVM Address'
        sheet['C1'] = 'Recipient'

        sheet['D1'] = 'Status'

        for cell in sheet._cells:
            sheet.cell(cell[0], cell[1]).font = Font(bold=True)

        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 46
        sheet.column_dimensions['C'].width = 46
        sheet.column_dimensions['D'].width = 20

        workbook.save('results/'+self.file_name)


    def edit_table(self, wallet: Wallet):
        while True:
            try:
                workbook = load_workbook('results/'+self.file_name)
                sheet = workbook.active

                valid_info = [
                    wallet.privatekey,
                    wallet.address,
                    wallet.recipient,
                    wallet.stats.get('status')
                ]
                sheet.append(valid_info)

                workbook.save('results/'+self.file_name)
                return True
            except PermissionError:
                logger.warning(f'Excel | Cant save excel file, close it!')
                sleep(3)
            except Exception as err:
                logger.critical(f'Excel | Cant save excel file: {err} | {wallet.address}')
                return False
